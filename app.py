import streamlit as st
import pandas as pd
import numpy as np
import easyocr
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from PIL import Image, ImageEnhance
import re
import os
import shutil
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

# Configuración de la página
st.set_page_config(page_title="📚 Inventario Biblioteca UCC - Sede Medellín", page_icon="📚", layout="wide")

# Colores para Excel
COLOR_VERDE = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
COLOR_MORADO = PatternFill(start_color="800080", end_color="800080", fill_type="solid")

# Ruta del Excel
EXCEL_PATH = "inventario.xlsx"
BACKUP_PATH = "inventario_backup.xlsx"

# Función para cargar OCR (cacheada)
@st.cache_resource
def cargar_ocr():
    return easyocr.Reader(['es', 'en'])

# Función para preprocesar imagen (mejora OCR)
def preprocesar_imagen(img):
    img_gray = img.convert('L')  # Escala de grises
    img_enhanced = ImageEnhance.Contrast(img_gray).enhance(2.0)  # Aumentar contraste
    return np.array(img_enhanced)

# Función para detectar códigos en texto
def detectar_codigos(textos):
    frases_prohibidas = [
        "sistemadeinformacionbibliografico",
        "sistemadeinformacion",
        "bibliografico",
        "biblioteca",
        "universidad",
        "cooperativa",
        "colombia"
    ]
    posibles_codigos = []
    for t in textos:
        t_limpio = t.lower().replace(" ", "").replace("-", "").strip()
        if any(frase in t_limpio for frase in frases_prohibidas):
            continue
        if re.fullmatch(r"b\d{6,8}", t_limpio):
            posibles_codigos.append(t_limpio.upper())
        elif t_limpio.startswith("b") and len(t_limpio) >= 7:
            posibles_codigos.append(t_limpio.upper())
    return max(posibles_codigos, key=len) if posibles_codigos else None

# Función para validar código
def validar_codigo(codigo, df):
    if not re.fullmatch(r"^B\d{6,8}$", codigo):
        return False, "Formato inválido (debe ser B seguido de 6-8 dígitos)."
    if codigo in df['codigo'].values:
        return False, "Código ya existe."
    return True, ""

# Función para actualizar Excel
def actualizar_excel(codigo, wb, sheet, codigo_a_fila, df):
    try:
        if codigo in codigo_a_fila:
            fila = codigo_a_fila[codigo]
            celda = f"A{fila}"
            sheet[celda].fill = COLOR_VERDE
            sheet[celda].font = Font(bold=True)
            return f"✔ Código {codigo} encontrado y marcado en verde."
        else:
            # Confirmación para nuevo código
            if st.button(f"Confirmar agregar código nuevo: {codigo}"):
                nueva_fila = sheet.max_row + 1
                sheet[f"A{nueva_fila}"] = codigo
                sheet[f"A{nueva_fila}"].fill = COLOR_MORADO
                sheet[f"A{nueva_fila}"].font = Font(bold=True)
                # Actualizar DataFrame en memoria
                nuevo_df = pd.concat([df, pd.DataFrame({'codigo': [codigo]})], ignore_index=True)
                st.session_state['df'] = nuevo_df
                return f"➕ Código nuevo agregado: {codigo}"
            return "Pendiente de confirmación."
    except Exception as e:
        return f"Error al actualizar Excel: {str(e)}"

# Función para crear backup
def crear_backup():
    if os.path.exists(EXCEL_PATH):
        shutil.copy(EXCEL_PATH, BACKUP_PATH)

# Función para exportar a PDF
def exportar_pdf(df, filename="inventario.pdf"):
    c = canvas.Canvas(filename, pagesize=letter)
    c.drawString(100, 750, "Inventario Biblioteca UCC")
    y = 720
    for index, row in df.iterrows():
        c.drawString(100, y, f"Código: {row['codigo']}")
        y -= 20
        if y < 50:
            c.showPage()
            y = 750
    c.save()

# Inicializar estado de sesión
if 'df' not in st.session_state:
    st.session_state['df'] = None
if 'codigos_detectados' not in st.session_state:
    st.session_state['codigos_detectados'] = []

# Título y descripción
st.title("📚 Inventario Biblioteca UCC - Sede Medellín")
with st.expander("📖 Guía de Uso"):
    st.write("""
    - **Escaneo**: Toma una foto clara del código de barras. La app lo detecta automáticamente.
    - **Manual**: Escribe el código si no puedes escanear.
    - **Batch**: Sube múltiples imágenes para procesar varias a la vez.
    - **Buscar**: Filtra el inventario por código.
    - Descarga el inventario actualizado en Excel, CSV o PDF.
    """)

# Sidebar para opciones
with st.sidebar:
    st.header("Opciones")
    if st.button("Crear Backup Manual"):
        crear_backup()
        st.success("Backup creado.")

# Cargar Excel inicial
if not os.path.exists(EXCEL_PATH):
    st.error("No se encontró 'inventario.xlsx'. Sube tu inventario inicial.")
    uploaded_file = st.file_uploader("Sube el inventario inicial", type=["xlsx"])
    if uploaded_file:
        with open(EXCEL_PATH, "wb") as f:
            f.write(uploaded_file.getbuffer())
        st.success("Inventario cargado. Recarga la página.")
    st.stop()

# Cargar datos
try:
    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active
    if st.session_state['df'] is None:
        st.session_state['df'] = pd.read_excel(EXCEL_PATH)
    df = st.session_state['df']
except Exception as e:
    st.error(f"Error al cargar Excel: {str(e)}")
    st.stop()

# Encontrar columna de códigos
codigo_columna = None
for col in df.columns:
    if "codigo" in col.lower():
        codigo_columna = col
        break
if not codigo_columna:
    st.error("No existe una columna 'codigo'.")
    st.stop()

codigo_a_fila = {str(row[codigo_columna]).strip(): idx + 2 for idx, row in df.iterrows()}

# Cargar OCR
reader = cargar_ocr()

# Sección de escaneo
st.subheader("Escanea el código")
img_file = st.camera_input("Toma una foto del código")
codigo_detectado = None

if img_file:
    with st.spinner("Procesando imagen..."):
        img = Image.open(img_file)
        img_array = preprocesar_imagen(img)
        textos = reader.readtext(img_array, detail=0)
        codigo_detectado = detectar_codigos(textos)
    
    if codigo_detectado:
        st.success(f"Código detectado: **{codigo_detectado}**")
        valido, mensaje = validar_codigo(codigo_detectado, df)
        if not valido:
            st.warning(mensaje)
        else:
            resultado = actualizar_excel(codigo_detectado, wb, sheet, codigo_a_fila, df)
            st.info(resultado)
            if "agregado" in resultado or "marcado" in resultado:
                crear_backup()
                wb.save(EXCEL_PATH)
                st.session_state['codigos_detectados'].append(codigo_detectado)
    else:
        st.warning("No se encontró un código válido. Intenta con una imagen más clara o ingresa manualmente.")

# Sección de batch (múltiples imágenes)
st.subheader("Procesar múltiples imágenes (Batch)")
uploaded_files = st.file_uploader("Sube imágenes", type=["jpg", "png", "jpeg"], accept_multiple_files=True)
if uploaded_files:
    for uploaded_file in uploaded_files:
        img = Image.open(uploaded_file)
        img_array = preprocesar_imagen(img)
        textos = reader.readtext(img_array, detail=0)
        codigo = detectar_codigos(textos)
        if codigo:
            valido, _ = validar_codigo(codigo, df)
            if valido:
                actualizar_excel(codigo, wb, sheet, codigo_a_fila, df)
                st.session_state['codigos_detectados'].append(codigo)
    crear_backup()
    wb.save(EXCEL_PATH)
    st.success("Batch procesado.")

# Sección manual
st.subheader("Ingresar código manualmente")
codigo_manual = st.text_input("Escribe el código:")
if codigo_manual:
    codigo_manual = codigo_manual.strip().upper()
    valido, mensaje = validar_codigo(codigo_manual, df)
    if not valido:
        st.warning(mensaje)
    else:
        resultado = actualizar_excel(codigo_manual, wb, sheet, codigo_a_fila, df)
        st.info(resultado)
        if "agregado" in resultado or "marcado" in resultado:
            crear_backup()
            wb.save(EXCEL_PATH)
            st.session_state['codigos_detectados'].append(codigo_manual)

# Sección de búsqueda
st.subheader("Buscar en inventario")
query = st.text_input("Buscar por código:")
if query:
    filtered_df = df[df[codigo_columna].str.contains(query, case=False, na=False)]
    st.dataframe(filtered_df)
else:
    st.dataframe(df)

# Descargas
st.subheader("Descargar inventario")
col1, col2, col3 = st.columns(3)
with col1:
    with open(EXCEL_PATH, "rb") as f:
        st.download_button("Descargar Excel", f, file_name="inventario_actualizado.xlsx")
with col2:
    csv = df.to_csv(index=False)
    st.download_button("Descargar CSV", csv, file_name="inventario.csv", mime="text/csv")
with col3:
    exportar_pdf(df)
    with open("inventario.pdf", "rb") as f:
        st.download_button("Descargar PDF", f, file_name="inventario.pdf")
